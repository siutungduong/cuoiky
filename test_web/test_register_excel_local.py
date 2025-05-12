import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

test_cases_file = "./test_cases_register.xlsx"
report_file = "./test_register_report.xlsx"

wb = openpyxl.load_workbook(test_cases_file)
ws = wb.active

header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
if "Actual Result" not in header_row:
    ws.cell(row=1, column=ws.max_column + 1, value="Actual Result")
    ws.cell(row=1, column=ws.max_column + 2, value="Status")

header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
actual_result_col = header_row.index("Actual Result") + 1
status_col = header_row.index("Status") + 1

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

signup_url = "http://localhost:3000/register"

for row in range(2, ws.max_row + 1):
    test_case = ws.cell(row=row, column=1).value
    first_name = ws.cell(row=row, column=2).value
    last_name = ws.cell(row=row, column=3).value
    email = ws.cell(row=row, column=4).value
    password = ws.cell(row=row, column=5).value
    confirm_password = ws.cell(row=row, column=6).value
    expected_result = ws.cell(row=row, column=7).value

    print(f"🔹 Đang kiểm thử: {test_case}")

    driver.get(signup_url)
    time.sleep(2)

    try:
        driver.find_element(By.XPATH, '//input[@placeholder="John"]').send_keys(first_name)
        driver.find_element(By.XPATH, '//input[@placeholder="Doe"]').send_keys(last_name)
        driver.find_element(By.XPATH, '//input[@placeholder="john@example.com"]').send_keys(email)

        password_fields = driver.find_elements(By.XPATH, '//input[@type="password" and @placeholder="********"]')
        password_fields[0].send_keys(password)
        password_fields[1].send_keys(confirm_password)

        driver.find_element(By.XPATH, '//button[contains(text(), "Đăng")]').click()

        time.sleep(3)

        if "login" in driver.current_url:
            actual_result = "Signed up successfully"
        else:
            try:
                driver.find_element(By.CLASS_NAME, "error-message")
                actual_result = "Signup failed"
            except:
                actual_result = "Signup failed"

        status = "Pass" if actual_result == expected_result else "Fail"
        ws.cell(row=row, column=actual_result_col, value=actual_result)
        ws.cell(row=row, column=status_col, value=status)

        print(f"✅ Expected: {expected_result} | Actual: {actual_result} | Status: {status}")

    except Exception as e:
        print(f"❌ Lỗi khi kiểm thử: {e}")
        ws.cell(row=row, column=actual_result_col, value="Test Failed")
        ws.cell(row=row, column=status_col, value="Fail")

wb.save(report_file)
wb.close()
driver.quit()
print(f"🎯 Hoàn thành kiểm thử! Báo cáo được lưu tại: {report_file}")
