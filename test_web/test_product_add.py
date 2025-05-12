import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

test_cases_file = "./product_test_cases.xlsx"
report_file = "./product_test_report.xlsx"

try:
    wb = openpyxl.load_workbook(test_cases_file)
    ws = wb.active
except FileNotFoundError:
    print(f"❌ Test cases file not found at: {test_cases_file}")
    exit(1)

headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

if "Actual Result" not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value="Actual Result")
    ws.cell(row=1, column=ws.max_column + 2, value="Status")
    actual_result_col = ws.max_column - 1
    status_col = ws.max_column
else:
    actual_result_col = headers.index("Actual Result") + 1
    status_col = headers.index("Status") + 1

try:
    col_type = headers.index("Type") + 1
    col_name = headers.index("Name") + 1
    col_price = headers.index("Price") + 1
    col_image = headers.index("Image") + 1
    col_brand = headers.index("Brand") + 1
    col_expected = headers.index("Expected Result") + 1
except ValueError as e:
    print(f"❌ Missing required column: {e}")
    wb.close()
    exit(1)

options = Options()
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

login_url = "http://localhost:3000/login"
add_product_url = "http://localhost:3000/admin/products/add"

# Login trước
try:
    print("🔑 Logging in...")
    driver.get(login_url)
    time.sleep(2)
    driver.find_element(By.ID, "email").send_keys("duc100100@gmail.com")  # chỉnh email
    driver.find_element(By.ID, "password").send_keys("123456")        # chỉnh password
    driver.find_element(By.CLASS_NAME, "login-button").click()
    time.sleep(3)
except Exception as e:
    print(f"❌ Login failed: {e}")
    driver.quit()
    wb.close()
    exit(1)

# Bắt đầu test các dòng trong Excel
for row in range(2, ws.max_row + 1):
    type_val = ws.cell(row=row, column=col_type).value or ""
    name_val = ws.cell(row=row, column=col_name).value or ""
    price_val = ws.cell(row=row, column=col_price).value or ""
    image_val = ws.cell(row=row, column=col_image).value or ""
    brand_val = ws.cell(row=row, column=col_brand).value or ""
    expected = ws.cell(row=row, column=col_expected).value

    print(f"🧪 Test: {name_val}...")

    try:
        driver.get(add_product_url)
        time.sleep(2)

        driver.find_element(By.NAME, "type").send_keys(type_val)
        driver.find_element(By.NAME, "name").send_keys(name_val)
        driver.find_element(By.NAME, "price").send_keys(str(price_val))
        driver.find_element(By.NAME, "image").send_keys(image_val)
        driver.find_element(By.NAME, "brand").send_keys(brand_val)

        driver.find_element(By.XPATH, "//button[contains(text(),'Thêm sản phẩm')]").click()
        time.sleep(2)

        actual = "Add failed"
        if "products" in driver.current_url:
            actual = "Add successful"

        status = "Pass" if actual == expected else "Fail"
        ws.cell(row=row, column=actual_result_col, value=actual)
        ws.cell(row=row, column=status_col, value=status)
        print(f"✅ Expected: {expected} | Actual: {actual} | Status: {status}")
    except Exception as e:
        print(f"❌ Error: {e}")
        ws.cell(row=row, column=actual_result_col, value="Test Failed")
        ws.cell(row=row, column=status_col, value="Fail")

wb.save(report_file)
wb.close()
driver.quit()
print(f"📊 Test complete. Report saved to: {report_file}")
