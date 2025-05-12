import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

test_cases_file = './Login_Test_Cases.xlsx'
report_file = './Login_Test_Report.xlsx'

wb = openpyxl.load_workbook(test_cases_file)
ws = wb.active

headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

if "Actual Result" not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value="Actual Result")
    ws.cell(row=1, column=ws.max_column + 2, value="Status")
    actual_result_col = ws.max_column - 1
    status_col = ws.max_column
else:
    actual_result_col = headers.index("Actual Result") + 1
    status_col = headers.index("Status") + 1

col_username = headers.index("Username") + 1
col_password = headers.index("Password") + 1
col_expected = headers.index("Expected Result") + 1

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

login_url = "http://localhost:3000/login"

for row in range(2, ws.max_row + 1):
    test_case = ws.cell(row=row, column=1).value
    username = ws.cell(row=row, column=col_username).value or ""
    password = ws.cell(row=row, column=col_password).value or ""
    expected_result = ws.cell(row=row, column=col_expected).value

    print(f"🔍 Đang kiểm thử: {test_case}")

    driver.get(login_url)
    time.sleep(2)

    try:
        email_field = driver.find_element(By.XPATH, "//input[@placeholder='john@example.com']")
        password_field = driver.find_element(By.XPATH, "//input[@placeholder='********']")
        login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Đăng nhập')]")

        email_field.clear()
        password_field.clear()
        email_field.send_keys(username)
        password_field.send_keys(password)
        login_button.click()

        time.sleep(3)

        if driver.current_url != login_url:
            actual_result = "Logged in successfully"
        else:
            try:
                driver.find_element(By.CLASS_NAME, "error-message")
                actual_result = "Login failed"
            except:
                actual_result = "Login failed"

        status = "Pass" if actual_result == expected_result else "Fail"

        ws.cell(row=row, column=actual_result_col, value=actual_result)
        ws.cell(row=row, column=status_col, value=status)

        print(f"✅ Expected: {expected_result} | Actual: {actual_result} | Status: {status}")

    except Exception as e:
        print(f"❌ Lỗi khi kiểm thử: {e}")
        ws.cell(row=row, column=actual_result_col, value="Test Failed")
        ws.cell(row=row, column=status_col, value="Fail")

wb.save(report_file)
wb.close()
driver.quit()

print(f"📊 Đã hoàn thành kiểm thử! Báo cáo được lưu tại: {report_file}")
